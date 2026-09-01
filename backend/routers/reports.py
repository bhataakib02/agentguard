import os
import io
import json
import csv
import datetime
from typing import Optional, List, Dict, Any
from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException, status, Header, Response
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from database import get_db
from core.deps import get_current_user, get_effective_org_id, check_org_isolation, HUMAN_ROLES
import models

# PDF & Excel Generation Engines
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/reports", tags=["Enterprise Report Center & Intelligence"])

STORAGE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "storage", "reports"))
os.makedirs(STORAGE_DIR, exist_ok=True)

class GenerateReportRequest(BaseModel):
    report_type: str  # EXECUTIVE, IAM, AGENTS, GOVERNANCE, SECURITY, AUDIT, FINANCIAL, LICENSE
    file_format: str = "PDF"  # PDF, EXCEL, CSV, TEXT
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    filters: Optional[Dict[str, Any]] = None

class ScheduleReportRequest(BaseModel):
    report_type: str
    file_format: str = "PDF"
    frequency: str = "WEEKLY"  # DAILY, WEEKLY, MONTHLY
    recipient_emails: str

REPORT_CATALOG = [
    {"id": "EXECUTIVE", "title": "Executive Summary Report", "category": "Executive & Compliance", "formats": ["PDF", "EXCEL"], "min_role": "ANALYST"},
    {"id": "ORGANIZATION", "title": "Organization Workspace Report", "category": "Executive & Compliance", "formats": ["PDF", "EXCEL"], "min_role": "ADMIN"},
    {"id": "IAM", "title": "User & IAM Security Report", "category": "IAM & Identity", "formats": ["EXCEL", "CSV", "PDF"], "min_role": "ADMIN"},
    {"id": "AGENTS", "title": "AI Agent Inventory Report", "category": "Agent Management", "formats": ["EXCEL", "CSV", "PDF"], "min_role": "OPERATOR"},
    {"id": "GOVERNANCE", "title": "Governance Decision Report", "category": "Governance & Control", "formats": ["PDF", "EXCEL", "CSV"], "min_role": "ANALYST"},
    {"id": "SECURITY", "title": "Security Incident & Threat Report", "category": "Security & SOC", "formats": ["PDF", "EXCEL", "CSV"], "min_role": "SECURITY_ANALYST"},
    {"id": "AUDIT", "title": "Audit Log & Provenance Report", "category": "Audit & Compliance", "formats": ["CSV", "EXCEL", "PDF"], "min_role": "ANALYST"},
    {"id": "FINANCIAL", "title": "Financial & Economics Report", "category": "Finance & Billing", "formats": ["EXCEL", "CSV", "PDF"], "min_role": "ADMIN"},
    {"id": "LICENSE", "title": "License Quota & Usage Report", "category": "Executive & Compliance", "formats": ["PDF", "EXCEL"], "min_role": "ADMIN"}
]

@router.get("")
def list_available_reports(current_user: models.User = Depends(get_current_user)):
    user_role = current_user.role
    allowed = []
    for r in REPORT_CATALOG:
        allowed.append(r)
    return allowed

@router.get("/history")
def get_report_history(
    current_user: models.User = Depends(get_current_user),
    x_org_context: Optional[str] = Header(None, alias="X-Organization-Context"),
    db: Session = Depends(get_db)
):
    target_org_id = get_effective_org_id(current_user, x_org_context)
    reports = db.query(models.ReportHistory).filter(
        models.ReportHistory.org_id == target_org_id
    ).order_by(models.ReportHistory.created_at.desc()).all()

    res = []
    for r in reports:
        res.append({
            "id": str(r.id),
            "report_type": r.report_type,
            "title": r.title,
            "file_format": r.file_format,
            "file_size_bytes": r.file_size_bytes,
            "created_at": r.created_at.isoformat() if r.created_at else None
        })
    return res

@router.get("/scheduled")
def get_scheduled_reports(
    current_user: models.User = Depends(get_current_user),
    x_org_context: Optional[str] = Header(None, alias="X-Organization-Context"),
    db: Session = Depends(get_db)
):
    target_org_id = get_effective_org_id(current_user, x_org_context)
    scheds = db.query(models.ScheduledReport).filter(
        models.ScheduledReport.org_id == target_org_id
    ).all()
    res = []
    for s in scheds:
        res.append({
            "id": str(s.id),
            "report_type": s.report_type,
            "title": s.title,
            "file_format": s.file_format,
            "frequency": s.frequency,
            "recipient_emails": s.recipient_emails,
            "status": s.status,
            "created_at": s.created_at.isoformat() if s.created_at else None
        })
    return res

@router.post("/schedule")
def create_scheduled_report(
    payload: ScheduleReportRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    sched = models.ScheduledReport(
        org_id=current_user.org_id,
        user_id=current_user.id,
        report_type=payload.report_type.upper(),
        title=f"Scheduled {payload.report_type.upper()} Report",
        file_format=payload.file_format.upper(),
        frequency=payload.frequency.upper(),
        recipient_emails=payload.recipient_emails,
        status="ACTIVE"
    )
    db.add(sched)
    db.commit()

    audit = models.AuditLog(
        event_type="REPORT_SCHEDULED",
        actor_type="USER",
        actor_id=str(current_user.id),
        action=f"Scheduled recurring {payload.frequency} report: {payload.report_type}",
        resource="reports",
        result="SUCCESS"
    )
    db.add(audit)
    db.commit()

    return {"status": "SUCCESS", "schedule_id": str(sched.id)}

@router.post("/generate")
def generate_report(
    payload: GenerateReportRequest,
    current_user: models.User = Depends(get_current_user),
    x_org_context: Optional[str] = Header(None, alias="X-Organization-Context"),
    db: Session = Depends(get_db)
):
    target_org_id = get_effective_org_id(current_user, x_org_context)
    check_org_isolation(current_user, target_org_id)

    org = db.query(models.Organization).filter(models.Organization.id == target_org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    lic = db.query(models.License).filter(models.License.org_id == target_org_id).first()

    org_dir = os.path.join(STORAGE_DIR, f"org_{target_org_id}")
    os.makedirs(org_dir, exist_ok=True)

    timestamp_str = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    report_type_clean = payload.report_type.upper()
    fmt_clean = payload.file_format.upper()

    filename = f"AGENTGUARD_{report_type_clean}_{timestamp_str}.{fmt_clean.lower() if fmt_clean != 'EXCEL' else 'xlsx'}"
    filepath = os.path.join(org_dir, filename)

    # 1. PDF Report Generation Engine
    if fmt_clean == "PDF":
        doc = SimpleDocTemplate(filepath, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            "DocTitle",
            parent=styles["Heading1"],
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#1F1F1F"),
            fontName="Times-Bold"
        )
        subtitle_style = ParagraphStyle(
            "SubTitle",
            parent=styles["Normal"],
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#2E9D50"),
            fontName="Times-Bold"
        )
        meta_style = ParagraphStyle(
            "Meta",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#666666"),
            fontName="Times-Roman"
        )
        heading2_style = ParagraphStyle(
            "DocH2",
            parent=styles["Heading2"],
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#8064C8"),
            fontName="Times-Bold",
            spaceBefore=12,
            spaceAfter=6
        )


        # Header Block
        elements.append(Paragraph("AGENTGUARD CONTROL PLANE", title_style))
        elements.append(Paragraph(f"CUSTOMER TENANT: {org.display_name or org.name} &nbsp;|&nbsp; licensed by AGENTGUARD", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#2E9D50"), spaceAfter=10, spaceBefore=4))

        meta_text = f"Report ID: RPT-{timestamp_str} &nbsp;|&nbsp; Generated: {datetime.datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}<br/>"
        meta_text += f"Generated By: {current_user.full_name} ({current_user.email}) &nbsp;|&nbsp; License Plan: {lic.plan_id if lic else 'STARTER'}"
        elements.append(Paragraph(meta_text, meta_style))
        elements.append(Spacer(1, 12))

        # Report Data Content
        if report_type_clean == "EXECUTIVE":
            elements.append(Paragraph("EXECUTIVE GOVERNANCE & SECURITY SUMMARY", heading2_style))
            user_count = db.query(models.User).filter(models.User.org_id == target_org_id, models.User.role != "SUPER_ADMIN").count()
            agent_count = db.query(models.Agent).filter(models.Agent.org_id == target_org_id).count()
            dec_count = db.query(models.Decision).join(models.Agent).filter(models.Agent.org_id == target_org_id).count()
            inc_count = db.query(models.SecurityIncident).filter(models.SecurityIncident.org_id == target_org_id).count()

            summary_data = [
                ["Metric Description", "Database Value", "Status / Governance State"],
                ["Total Organization Users", str(user_count), "ACTIVE"],
                ["Autonomous AI Agents", str(agent_count), "MONITORED"],
                ["Decision Black Box Evaluations", str(dec_count), "AUDITED"],
                ["Security Incidents Logged", str(inc_count), "ISOLATED" if inc_count == 0 else "WARNING"],
                ["Tenant License Status", lic.status if lic else "ACTIVE", "ENFORCED"]
            ]
            t = Table(summary_data, colWidths=[240, 120, 180])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#F1EDFA")),
                ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#8064C8")),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E8E8E4")),
                ('PADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(t)

        elif report_type_clean == "SECURITY":
            elements.append(Paragraph("SECURITY INCIDENTS & THREAT EVALUATION", heading2_style))
            incidents = db.query(models.SecurityIncident).filter(models.SecurityIncident.org_id == target_org_id).all()
            sec_data = [["Incident ID", "Title", "Severity", "Status", "Timestamp"]]
            for inc in incidents:
                sec_data.append([str(inc.id)[:8], inc.title[:30], inc.severity, inc.status, inc.timestamp.strftime('%Y-%m-%d %H:%M')])
            if len(sec_data) == 1:
                sec_data.append(["N/A", "No Active Security Incidents Logged", "LOW", "CLEAN", "-"] )

            t = Table(sec_data, colWidths=[80, 220, 80, 80, 80])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#FDECEC")),
                ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#E53935")),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E8E8E4")),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            elements.append(t)

        else:
            elements.append(Paragraph(f"{report_type_clean} SUBSYSTEM DETAILED REPORT", heading2_style))
            gen_data = [
                ["Field", "Value"],
                ["Organization Name", org.name],
                ["Tenant Domain", org.domain or "N/A"],
                ["License Plan", lic.plan_id if lic else "STARTER"],
                ["Report Type", report_type_clean],
                ["Report Scope", "TENANT_ISOLATED"]
            ]
            t = Table(gen_data, colWidths=[200, 340])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#EAF7EE")),
                ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#237A3C")),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E8E8E4")),
                ('PADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(t)

        # Footer
        elements.append(Spacer(1, 20))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#E8E8E4"), spaceBefore=10))
        elements.append(Paragraph("AGENTGUARD — Runtime Control Plane for Autonomous AI &nbsp;|&nbsp; licensed by AGENTGUARD", meta_style))

        doc.build(elements)

    # 2. Excel Generation Engine
    elif fmt_clean in ["EXCEL", "XLSX"]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report Summary"

        ws.append(["AGENTGUARD CONTROL PLANE — ENTERPRISE REPORT"])
        ws.append(["Organization:", org.name, "Plan:", lic.plan_id if lic else "STARTER"])
        ws.append(["Report Type:", report_type_clean, "Generated:", datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')])
        ws.append([])

        if report_type_clean == "IAM":
            ws.append(["User ID", "Email", "Full Name", "Role", "Department", "Status"])
            users = db.query(models.User).filter(models.User.org_id == target_org_id, models.User.role != "SUPER_ADMIN").all()
            for u in users:
                ws.append([str(u.id), u.email, u.full_name, u.role, u.department or "General", u.status])
        elif report_type_clean == "AGENTS":
            ws.append(["Agent ID", "Name", "Role", "Autonomy Level", "Risk Score", "Status"])
            agents = db.query(models.Agent).filter(models.Agent.org_id == target_org_id).all()
            for a in agents:
                ws.append([str(a.id), a.name, a.role, a.autonomy_level, a.risk_score, a.status])
        else:
            ws.append(["Metric", "Value"])
            ws.append(["Total Users", db.query(models.User).filter(models.User.org_id == target_org_id, models.User.role != "SUPER_ADMIN").count()])
            ws.append(["Total Agents", db.query(models.Agent).filter(models.Agent.org_id == target_org_id).count()])

        wb.save(filepath)

    # 3. CSV Generation Engine
    else:
        with open(filepath, mode="w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["AGENTGUARD CONTROL PLANE REPORT", report_type_clean])
            writer.writerow(["Organization", org.name, "Domain", org.domain or ""])
            writer.writerow(["Generated At", datetime.datetime.utcnow().isoformat()])
            writer.writerow([])

            if report_type_clean == "IAM":
                writer.writerow(["User ID", "Email", "Full Name", "Role", "Department", "Status"])
                users = db.query(models.User).filter(models.User.org_id == target_org_id, models.User.role != "SUPER_ADMIN").all()
                for u in users:
                    writer.writerow([str(u.id), u.email, u.full_name, u.role, u.department or "General", u.status])
            else:
                writer.writerow(["Metric", "Value"])
                writer.writerow(["Organization Name", org.name])
                writer.writerow(["Total Users", db.query(models.User).filter(models.User.org_id == target_org_id, models.User.role != "SUPER_ADMIN").count()])
                writer.writerow(["Total Agents", db.query(models.Agent).filter(models.Agent.org_id == target_org_id).count()])

    file_size = os.path.getsize(filepath)

    rec = models.ReportHistory(
        org_id=target_org_id,
        user_id=current_user.id,
        report_type=report_type_clean,
        title=f"{report_type_clean.capitalize()} Report",
        file_format=fmt_clean,
        file_path=filepath,
        file_size_bytes=file_size,
        filters_json=payload.filters or {}
    )
    db.add(rec)
    db.commit()

    audit = models.AuditLog(
        event_type="REPORT_GENERATED",
        actor_type="USER",
        actor_id=str(current_user.id),
        action=f"Generated {fmt_clean} report for {report_type_clean}",
        resource="reports",
        result="SUCCESS",
        metadata_json={"report_id": str(rec.id), "format": fmt_clean}
    )
    db.add(audit)
    db.commit()

    return {
        "status": "SUCCESS",
        "report_id": str(rec.id),
        "title": rec.title,
        "file_format": fmt_clean,
        "file_size_bytes": file_size,
        "filename": filename
    }

@router.get("/download/{report_id}")
def download_report(
    report_id: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    rec = db.query(models.ReportHistory).filter(models.ReportHistory.id == report_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Report record not found")

    check_org_isolation(current_user, str(rec.org_id))

    if not os.path.exists(rec.file_path):
        raise HTTPException(status_code=404, detail="Report file not found on disk")

    audit = models.AuditLog(
        event_type="REPORT_DOWNLOADED",
        actor_type="USER",
        actor_id=str(current_user.id),
        action=f"Downloaded report {rec.title} ({rec.file_format})",
        resource="reports",
        result="SUCCESS",
        metadata_json={"report_id": str(rec.id)}
    )
    db.add(audit)
    db.commit()

    return FileResponse(rec.file_path, filename=os.path.basename(rec.file_path))
